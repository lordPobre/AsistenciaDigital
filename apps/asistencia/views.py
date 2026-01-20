import base64
import openpyxl
import hashlib
import datetime
import calendar
import json
import google.generativeai as genai
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import models
from django.db import transaction
from django.db.models import Min, Max, Count
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta, time
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from weasyprint import HTML
from django.core.exceptions import ValidationError
from geopy.geocoders import Nominatim
from django.views.decorators.cache import cache_control
from django.views.generic import TemplateView
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.html import strip_tags
from .models import Marcacion, Empresa, SolicitudMarca, Feriado, Vacacion, LicenciaMedica, Perfil, DiaAdministrativo
from .forms import VacacionForm, LicenciaForm



# =======================================================
# 1. FUNCIONES DE PERMISOS Y ROLES
# =======================================================

def es_fiscalizador(user):
    return user.is_authenticated and hasattr(user, 'perfil') and user.perfil.rol == 'FISCALIZADOR'

def es_empleador(user):
    # Consideramos empleador si tiene el rol o es Superusuario
    if user.is_superuser: return True
    return user.is_authenticated and hasattr(user, 'perfil') and user.perfil.rol == 'EMPLEADOR'

def es_trabajador(user):
    return user.is_authenticated and hasattr(user, 'perfil') and user.perfil.rol == 'TRABAJADOR'

genai.configure(api_key="AIzaSyBuRNjmjsi5AKSWO9g6ggGLtrJKxuKbty0")

class ServiceWorkerView(TemplateView):
    template_name = "asistencia/sw.js"
    content_type = "application/javascript"


# =======================================================
# 2. VISTAS PRINCIPALES (HOME Y MARCAJE)
# =======================================================

@login_required
def home(request):
    # 1. Chequeo de cambio de contraseña obligatorio
    if hasattr(request.user, 'perfil') and request.user.perfil.cambiar_pass_inicial:
        return redirect('cambiar_password_obligatorio')

    # 2. Cargar datos para el Dashboard
    ultimas_marcas = Marcacion.objects.filter(trabajador=request.user).order_by('-timestamp')[:5]
    ultima_marca = Marcacion.objects.filter(trabajador=request.user).order_by('-timestamp').first()
    solicitudes_pendientes = SolicitudMarca.objects.filter(trabajador=request.user, estado='PENDIENTE').exclude(solicitante=request.user) # <--- ESTO HACE LA MAGIA

    contexto = {
        'marcas': ultimas_marcas,
        'ultima_marca': ultima_marca,
        'solicitudes': solicitudes_pendientes,
    }

    return render(request, 'asistencia/dashboard.html', contexto)

@login_required
def registrar_marca(request):
    # Validamos que sea POST
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    # =================================================================
    # 1. OBTENCIÓN DE DATOS HÍBRIDA (Fuera del Try Principal)
    # =================================================================
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        data = request.POST

    # Extraemos variables
    tipo = data.get('tipo', 'ENTRADA')
    raw_lat = data.get('latitud')
    raw_lon = data.get('longitud')
    foto_b64 = data.get('foto_base64')
    ip = request.META.get('REMOTE_ADDR')
    
    # Datos extra
    animo_recibido = data.get('animo')
    comentario_recibido = data.get('comentario_animo')
    fecha_offline_str = data.get('fecha_offline')

    # Procesar fecha (si viene de offline)
    timestamp_real = timezone.now()
    if fecha_offline_str:
        try:
            timestamp_real = datetime.fromisoformat(fecha_offline_str.replace('Z', '+00:00'))
        except ValueError:
            pass

    # =================================================================
    # INICIO DEL TRY PRINCIPAL (Aquí empieza la protección lógica)
    # =================================================================
    try:
        # --- 2. PROCESAR GPS ---
        lat, lon = 0, 0
        if raw_lat and str(raw_lat).lower() != 'nan':
            try:
                lat = "{:.7f}".format(float(raw_lat))
                lon = "{:.7f}".format(float(raw_lon))
            except (ValueError, TypeError):
                lat, lon = 0, 0

        # --- 3. OBTENER DIRECCIÓN (GEOCODING) ---
        direccion_texto = "Ubicación no detectada"
        if lat and lon:
            try:
                geolocator = Nominatim(user_agent="asistencia_perseus_v1", timeout=5)
                location = geolocator.reverse(f"{lat}, {lon}", timeout=5)
                if location:
                        direccion_texto = location.address
            except Exception as e:
                print(f"Error Geopy: {e}")
                direccion_texto = "Error de conexión con mapa"

        # --- 4. CREAR INSTANCIA ---
        nueva_marca = Marcacion(
            trabajador=request.user,
            tipo=tipo,
            latitud=lat,
            longitud=lon,
            timestamp=timestamp_real,
            direccion=direccion_texto,
            ip_address=ip,
            animo=animo_recibido,
            comentario_animo=comentario_recibido
        )

        # --- 5. PROCESAR FOTO (Cloudinary) ---
        if foto_b64:
            try:
                if ";base64," in foto_b64:
                    format_data, imgstr = foto_b64.split(';base64,')
                    ext = format_data.split('/')[-1]
                    if not ext: ext = "jpg"
                else:
                    imgstr = foto_b64
                    ext = "jpg"

                nombre_foto = f'marca_{request.user.id}_{int(timezone.now().timestamp())}.{ext}'
                archivo_imagen = ContentFile(base64.b64decode(imgstr), name=nombre_foto)
                
                # Asignamos al modelo
                nueva_marca.foto = archivo_imagen
            except Exception as e:
                print(f"Error procesando foto: {e}")
                return JsonResponse({'error': 'Error al procesar la imagen.'}, status=400)

        # --- 6. VALIDACIONES HARDWARE ---
        if tipo in ['ENTRADA', 'SALIDA']:
            if float(nueva_marca.latitud) == 0:
                return JsonResponse({'error': 'Hardware: GPS no detectado.'}, status=400)
            if not nueva_marca.foto:
                return JsonResponse({'error': 'Hardware: Foto no detectada.'}, status=400)

        # --- 7. GUARDAR (Sube a Cloudinary y guarda en BD) ---
        nueva_marca.save()

        # --- 8. ENVÍO DE CORREO PROFESIONAL (HTML) ---
        if request.user.email:
            try:
                # A. Preparar datos
                hora_fmt = timezone.localtime(nueva_marca.timestamp).strftime('%H:%M:%S')
                fecha_fmt = timezone.localtime(nueva_marca.timestamp).strftime('%d/%m/%Y')

                # Obtener datos de la empresa
                datos_empresa = Empresa.objects.first()
                nombre_empresa = datos_empresa.nombre if datos_empresa else "Su Empresa"
                email_rrhh = datos_empresa.email_rrhh if datos_empresa else None

                # Enlace a Google Maps
                link_maps = f"https://www.google.com/maps?q={nueva_marca.latitud},{nueva_marca.longitud}"
                ubicacion_texto = nueva_marca.direccion if nueva_marca.direccion else "Coordenadas GPS"

                # B. Definir Asunto
                asunto = f'✅ Comprobante de Asistencia: {tipo} - {request.user.get_full_name()}'

                # C. Crear el Mensaje en HTML (TU DISEÑO EXACTO)
                html_message = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <style>
                            body {{ font-family: Arial, sans-serif; color: #333333; }}
                            .container {{ max-width: 600px; margin: 0 auto; border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden; }}
                            .header {{ background-color: #004085; color: #ffffff; padding: 20px; text-align: center; }}
                            .content {{ padding: 25px; background-color: #ffffff; }}
                            .detail-table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 20px; }}
                            .detail-table td {{ padding: 10px; border-bottom: 1px solid #f0f0f0; }}
                            .label {{ font-weight: bold; color: #555555; width: 40%; }}
                            .footer {{ background-color: #f8f9fa; padding: 15px; text-align: center; font-size: 12px; color: #888888; border-top: 1px solid #e0e0e0; }}
                            .btn {{ display: inline-block; padding: 8px 12px; background-color: #28a745; color: white; text-decoration: none; border-radius: 4px; font-size: 12px; }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <h2 style="margin:0;">Registro de Asistencia</h2>
                                <p style="margin:5px 0 0; font-size: 14px; opacity: 0.9;">{nombre_empresa}</p>
                            </div>
                            <div class="content">
                                <p>Estimado/a <strong>{request.user.first_name} {request.user.last_name}</strong>,</p>
                                <p>El sistema ha procesado exitosamente su marcación. A continuación se detallan los datos del registro:</p>

                                <table class="detail-table">
                                    <tr>
                                        <td class="label">Tipo de Marca:</td>
                                        <td><strong style="color: #004085;">{tipo}</strong></td>
                                    </tr>
                                    <tr>
                                        <td class="label">Fecha:</td>
                                        <td>{fecha_fmt}</td>
                                    </tr>
                                    <tr>
                                        <td class="label">Hora Registrada:</td>
                                        <td>{hora_fmt}</td>
                                    </tr>
                                    <tr>
                                        <td class="label">Ubicación:</td>
                                        <td>
                                            {ubicacion_texto}<br>
                                            <a href="{link_maps}" class="btn" style="color: white; margin-top:5px;">Ver en Mapa</a>
                                        </td>
                                    </tr>
                                    <tr>
                                        <td class="label">Estado:</td>
                                        <td><span style="color:green;">✔ Validado Exitosamente</span></td>
                                    </tr>
                                </table>

                                <p style="font-size: 13px; color: #666;">Este registro ha sido almacenado en nuestra base de datos segura y servirá como respaldo oficial de su jornada laboral.</p>
                            </div>
                            <div class="footer">
                                <p>Este es un mensaje automático generado por el Sistema de Gestión de Asistencia de {nombre_empresa}.<br>
                                Por favor, no responda a este correo.</p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """

                # D. Versión Texto Plano y Remitente
                plain_message = strip_tags(html_message)
                nombre_mostrar = "Sistema de Asistencia"
                # Usamos el remitente configurado en settings
                remitente = f"{nombre_mostrar} <{settings.EMAIL_HOST_USER}>"

                # E. Configurar Destinatarios (AQUÍ ESTABA EL ERROR DE INDENTACIÓN CORREGIDO)
                destinatarios = [request.user.email]
                if email_rrhh:
                    destinatarios.append(email_rrhh)
                    print(f"✅ Copia enviada a RRHH: {email_rrhh}")

                # F. Enviar (AHORA FUERA DEL IF DE RRHH)
                send_mail(
                    subject=asunto,
                    message=plain_message,
                    from_email=remitente, # Usa settings para evitar error de sender
                    recipient_list=destinatarios,
                    html_message=html_message,
                    fail_silently=False # <--- Activado para ver errores
                )
                print("✅ Correo enviado exitosamente.")

            except Exception as e:
                print(f"❌ Error enviando correo: {e}")

        # --- 9. RESPUESTA FINAL EXITOSA ---
        return JsonResponse({'status': 'ok', 'mensaje': 'Marca registrada correctamente.'})

    # =================================================================
    # CIERRE DEL TRY PRINCIPAL (Los excepts van aquí, alineados con el try)
    # =================================================================
    except Exception as e:
        print(f"Error crítico en servidor: {e}")
        return JsonResponse({'error': f"Error del sistema: {str(e)}"}, status=500)
    

@login_required
def mis_marcas(request):
    """Vista para que el trabajador vea su propio historial."""
    marcas = Marcacion.objects.filter(trabajador=request.user).order_by('-timestamp')
    return render(request, 'asistencia/mis_marcas.html', {'marcas': marcas})


# =======================================================
# 3. PANELES DE GESTIÓN (EMPLEADOR Y FISCALIZADOR)
# =======================================================

@login_required
@user_passes_test(es_empleador)
def panel_empresa(request):
    """Panel para el Jefe/RRHH con filtros y buscador."""
    try:
        mi_empresa = request.user.perfil.empresa
    except AttributeError:
        messages.error(request, "Su usuario no tiene empresa asignada.")
        return redirect('home')

    marcas = Marcacion.objects.filter(trabajador__perfil__empresa=mi_empresa).order_by('-timestamp')

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    busqueda = request.GET.get('busqueda')

    if fecha_inicio and fecha_fin:
        marcas = marcas.filter(timestamp__date__range=[fecha_inicio, fecha_fin])

    if busqueda:
        marcas = marcas.filter(
            Q(trabajador__first_name__icontains=busqueda) |
            Q(trabajador__last_name__icontains=busqueda) |
            Q(trabajador__perfil__rut__icontains=busqueda)
        )

    context = {
        'marcas': marcas,
        'nombre_empresa': mi_empresa.nombre,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'busqueda': busqueda
    }
    return render(request, 'asistencia/panel_empresa.html', context)

@staff_member_required
def panel_rrhh(request):
    # 1. LOGICA ORIGINAL: Buscamos solicitudes de marcas
    solicitudes_trabajadores = SolicitudMarca.objects.filter(
        estado='PENDIENTE'
    ).filter(
        solicitante=models.F('trabajador')
    ).order_by('created_at')

    # 2. NUEVA LÓGICA: Buscamos solicitudes de VACACIONES 🌴
    # (Esto es lo que faltaba para llenar la segunda tabla del HTML)
    vacaciones_pendientes = Vacacion.objects.filter(
        estado='PENDIENTE'
    ).order_by('inicio')

    # --- 3. NUEVO: Solicitudes de Días Administrativos ---
    dias_admin_pendientes = DiaAdministrativo.objects.filter(
        estado='PENDIENTE'
    ).order_by('fecha')
    # -----------------------------------------------------

    # 1. FILTRO: Últimos 7 días
    fecha_inicio = timezone.now() - timedelta(days=7)

    # 2. CONSULTA AGRUPADA (La magia de Django)
    # Esto devuelve algo como: [{'animo': 'FELIZ', 'total': 15}, {'animo': 'MOLESTO', 'total': 3}]
    data_animo = Marcacion.objects.filter(
        timestamp__gte=fecha_inicio,
        animo__isnull=False
    ).values('animo').annotate(total=Count('animo'))

    # 3. PROCESAR PARA EL GRÁFICO
    # Inicializamos contadores en 0 por si nadie ha marcado alguna emoción
    conteo = {'FELIZ': 0, 'NEUTRAL': 0, 'MOLESTO': 0}

    for item in data_animo:
        if item['animo'] in conteo:
            conteo[item['animo']] = item['total']

    # Calculamos el total para sacar porcentajes si quisieras, o pasamos los números directos
    context = {
        'solicitudes': solicitudes_trabajadores,       # Para la tabla de marcas
        'vacaciones_pendientes': vacaciones_pendientes,
        'dias_admin_pendientes': dias_admin_pendientes,
        'grafico_feliz': conteo['FELIZ'],
        'grafico_neutral': conteo['NEUTRAL'],
        'grafico_molesto': conteo['MOLESTO'],
    }

    # 3. Enviamos ambas listas al template
    return render(request, 'asistencia/panel_rrhh.html', context)

@login_required
@user_passes_test(es_fiscalizador)
def panel_fiscalizador(request):
    """
    Panel Auditoría DT con cálculo de jornadas (Entrada vs Salida).
    """
    # 1. SEGURIDAD: Obtener empresa del fiscalizador logueado
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.empresa:
        messages.error(request, "Usuario fiscalizador sin empresa asignada.")
        return redirect('home')

    empresa = perfil.empresa

    # 2. FILTROS DE FECHA (Vital para que no colapse el sistema)
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    # Base de consulta: Solo marcas de ESTA empresa, ordenadas para el algoritmo
    # select_related optimiza la consulta a la BD
    marcas_qs = Marcacion.objects.filter(
        trabajador__perfil__empresa=empresa
    ).select_related('trabajador', 'trabajador__perfil').order_by('trabajador', 'timestamp')

    if desde and hasta:
        # Agregamos un día extra al final para asegurar cobertura completa de fechas
        marcas_qs = marcas_qs.filter(timestamp__range=[desde, hasta + " 23:59:59"])
    else:
        # Por defecto: Últimos 30 días si no hay filtro (para no cargar años de historia)
        mes_atras = timezone.now() - datetime.timedelta(days=30)
        marcas_qs = marcas_qs.filter(timestamp__gte=mes_atras)

    # 3. ALGORITMO DE EMPAREJAMIENTO (Tu lógica, optimizada) 🧠
    jornadas = []
    entrada_pendiente = None

    # Iteramos sobre el QuerySet ya filtrado
    for marca in marcas_qs:
        # Si cambiamos de trabajador, reiniciamos el ciclo (por seguridad)
        if entrada_pendiente and entrada_pendiente.trabajador != marca.trabajador:
             # Cerramos la anterior como inconclusa
             jornadas.append({
                'trabajador': entrada_pendiente.trabajador,
                'fecha': entrada_pendiente.timestamp,
                'entrada': entrada_pendiente.timestamp,
                'salida': None,
                'duracion': 'Sin marcación de salida',
                'estado': 'warning' # Amarillo
            })
             entrada_pendiente = None

        if marca.tipo == 'ENTRADA':
            if entrada_pendiente:
                # Doble entrada (Error)
                jornadas.append({
                    'trabajador': entrada_pendiente.trabajador,
                    'fecha': entrada_pendiente.timestamp,
                    'entrada': entrada_pendiente.timestamp,
                    'salida': None,
                    'duracion': 'Error: Doble Entrada',
                    'estado': 'warning'
                })
            entrada_pendiente = marca

        elif marca.tipo == 'SALIDA':
            if entrada_pendiente:
                # CICLO CORRECTO: Entrada + Salida
                diferencia = marca.timestamp - entrada_pendiente.timestamp
                total_segundos = int(diferencia.total_seconds())
                horas = total_segundos // 3600
                minutos = (total_segundos % 3600) // 60

                jornadas.append({
                    'trabajador': marca.trabajador,
                    'fecha': marca.timestamp,
                    'entrada': entrada_pendiente.timestamp,
                    'salida': marca.timestamp,
                    'duracion': f"{horas}h {minutos}m",
                    'estado': 'success' # Verde
                })
                entrada_pendiente = None
            else:
                # Salida huérfana (Sin entrada previa)
                jornadas.append({
                    'trabajador': marca.trabajador,
                    'fecha': marca.timestamp,
                    'entrada': None,
                    'salida': marca.timestamp,
                    'duracion': 'Error: Falta Entrada',
                    'estado': 'danger' # Rojo
                })

        # Ignoramos Colaciones por ahora para simplificar el reporte DT

    # Check final por si quedó alguno colgado
    if entrada_pendiente:
        jornadas.append({
            'trabajador': entrada_pendiente.trabajador,
            'fecha': entrada_pendiente.timestamp,
            'entrada': entrada_pendiente.timestamp,
            'salida': None,
            'duracion': 'En curso (Trabajando)',
            'estado': 'info' # Azul
        })

    # Invertimos para ver lo más reciente arriba
    jornadas.reverse()

    context = {
        'empresa': empresa,
        'jornadas': jornadas,
        'desde': desde,
        'hasta': hasta
    }
    return render(request, 'asistencia/panel_fiscalizador.html', context)


# =======================================================
# 4. REPORTES Y EXPORTACIÓN
# =======================================================

def exportar_excel_empresa(request):
    """
    Genera el reporte interno para RRHH (Imagen: Detalle de marcas, colación y horas).
    NO incluye Hash. Es para gestión.
    """
    # 1. Validar Empresa
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.empresa:
        return redirect('home')

    # 2. Obtener marcas de la empresa
    marcas = Marcacion.objects.filter(trabajador__perfil__empresa=perfil.empresa).order_by('trabajador', 'timestamp')

    # 3. Filtros de Fecha (Opcional)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio and fecha_fin:
        try:
            marcas = marcas.filter(timestamp__date__range=[fecha_inicio, fecha_fin])
        except: pass

    # 4. Procesar Datos (Agrupar por Día y Persona)
    reporte = {}
    for marca in marcas:
        fecha_local = timezone.localtime(marca.timestamp)
        # Clave única: ID Trabajador + Fecha
        key = (marca.trabajador.id, fecha_local.date())

        if key not in reporte:
            p = getattr(marca.trabajador, 'perfil', None)
            reporte[key] = {
                'fecha': fecha_local.date(),
                'empresa': perfil.empresa.nombre,
                'trabajador': f"{marca.trabajador.first_name} {marca.trabajador.last_name}",
                'rut': p.rut if p else "S/I",
                'cargo': p.cargo if p else "S/I",
                'entrada': None, 'inicio_col': None, 'fin_col': None, 'salida': None
            }

        hora = fecha_local.time()
        if marca.tipo == 'ENTRADA':
            if reporte[key]['entrada'] is None or hora < reporte[key]['entrada']: reporte[key]['entrada'] = hora
        elif marca.tipo == 'INICIO_COLACION': reporte[key]['inicio_col'] = hora
        elif marca.tipo == 'FIN_COLACION': reporte[key]['fin_col'] = hora
        elif marca.tipo == 'SALIDA':
             if reporte[key]['salida'] is None or hora > reporte[key]['salida']: reporte[key]['salida'] = hora

    # 5. Crear Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Detallado_{perfil.empresa.nombre}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Asistencia"

    # Encabezados (IDÉNTICOS A TU FOTO)
    headers = ["Fecha", "Empresa", "Trabajador", "RUT", "Cargo", "Entrada", "Ini Col", "Fin Col", "Tiempo Col.", "Salida", "Horas Trab"]
    ws.append(headers)

    # Estilo Encabezado
    fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    # 6. Llenar Filas y Calcular
    for data in reporte.values():
        entrada, salida = data['entrada'], data['salida']
        inicio_col, fin_col = data['inicio_col'], data['fin_col']

        tiempo_col_str = ""
        horas_trab_str = ""

        # A. Calcular Tiempo Colación
        if inicio_col and fin_col:
            dummy = datetime.min
            diff = datetime.combine(dummy, fin_col) - datetime.combine(dummy, inicio_col)
            segundos = diff.total_seconds()
            tiempo_col_str = f"{int(segundos//3600):02d}:{int((segundos%3600)//60):02d}"

        # B. Calcular Horas Trabajadas (Bruto: Salida - Entrada)
        # Nota: Si quieres restar colación automáticamente, descomenta la lógica de resta
        if entrada and salida:
            dummy = datetime.min
            diff = datetime.combine(dummy, salida) - datetime.combine(dummy, entrada)
            segundos = diff.total_seconds()
            horas_trab_str = f"{int(segundos//3600):02d}:{int((segundos%3600)//60):02d}"

        ws.append([
            data['fecha'].strftime("%d/%m/%Y"),
            data['empresa'],
            data['trabajador'],
            data['rut'],
            data['cargo'],
            entrada.strftime("%H:%M") if entrada else "--",
            inicio_col.strftime("%H:%M") if inicio_col else "--",
            fin_col.strftime("%H:%M") if fin_col else "--",
            tiempo_col_str, # Columna Nueva
            salida.strftime("%H:%M") if salida else "--",
            horas_trab_str  # Columna Final
        ])

    # Auto-ancho
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    ws.column_dimensions['C'].width = 25 # Nombre más ancho

    wb.save(response)
    return response

def exportar_clima_laboral(request):
    # 1. Crear el libro de Excel y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clima Laboral"

    # 2. Definir los Encabezados de la tabla
    headers = ['Fecha', 'Hora', 'Trabajador', 'RUT', 'Cargo', 'Estado de Ánimo', 'Comentario / Motivo']
    ws.append(headers)

    # 3. Estilo para el encabezado (Negrita, fondo gris, centrado)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

    for cell in ws[1]:  # Fila 1
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 4. Obtener solo las marcas de SALIDA que tengan algún ánimo registrado
    marcas = Marcacion.objects.filter(tipo='SALIDA', animo__isnull=False).order_by('-timestamp')

    # 5. Escribir los datos
    for marca in marcas:
        # Convertir fecha y hora para que se vean bien
        fecha = marca.timestamp.strftime("%d/%m/%Y")
        hora = marca.timestamp.strftime("%H:%M")

        # Poner un Emoji en el Excel según el ánimo
        animo_texto = marca.get_animo_display() # Obtiene "Feliz" en vez de "FELIZ"
        if marca.animo == 'FELIZ':
            animo_display = f"😆 {animo_texto}"
        elif marca.animo == 'NEUTRAL':
            animo_display = f"😐 {animo_texto}"
        elif marca.animo == 'MOLESTO':
            animo_display = f"😫 {animo_texto}"
        else:
            animo_display = animo_texto

        # Si no hay comentario, poner guion
        comentario = marca.comentario_animo if marca.comentario_animo else "-"

        # Agregar la fila
        ws.append([
            fecha,
            hora,
            f"{marca.trabajador.first_name} {marca.trabajador.last_name}",
            marca.trabajador.perfil.rut,
            marca.trabajador.perfil.cargo,
            animo_display,
            comentario
        ])

    # 6. Ajustar ancho de columnas automáticamente (estético)
    ws.column_dimensions['A'].width = 12 # Fecha
    ws.column_dimensions['C'].width = 25 # Nombre
    ws.column_dimensions['E'].width = 20 # Cargo
    ws.column_dimensions['F'].width = 15 # Animo
    ws.column_dimensions['G'].width = 50 # Comentario (ancho para leer bien)

    # 7. Preparar la respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Clima_Laboral.xlsx"'

    wb.save(response)
    return response

@login_required
def generar_pdf(request):
    """Genera comprobante PDF para el trabajador."""
    timezone.activate(timezone.get_current_timezone())
    marcas = Marcacion.objects.filter(trabajador=request.user).order_by('timestamp')
    empresa = request.user.perfil.empresa
    logo_path = None
    if empresa and empresa.logo:
        # 1. Obtenemos la ruta física
        ruta_absoluta = empresa.logo.path
        logo_path = f"file://{ruta_absoluta}"
    contexto = {
        'marcas': marcas,
        'usuario': request.user,
        'fecha_generacion': timezone.localtime(timezone.now()),
        'empresa': empresa,
        'logo_path': logo_path,
    }

    html_string = render_to_string('reportes/libro_asistencia.html', contexto)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="asistencia_{request.user.username}.pdf"'
    HTML(string=html_string).write_pdf(response)
    return response


# =======================================================
# 5. UTILIDADES (CAMBIO CLAVE, AYUDA)
# =======================================================

@login_required
def cambiar_password_obligatorio(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            if hasattr(user, 'perfil'):
                user.perfil.cambiar_pass_inicial = False
                user.perfil.save()
            messages.success(request, '¡Contraseña actualizada!')
            return redirect('home')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'asistencia/cambiar_pass.html', {'form': form})

@login_required
def manual_ayuda(request):
    return render(request, 'asistencia/ayuda.html')

@login_required
def gestionar_solicitudes(request):
    # Listamos solo las pendientes para este trabajador
    solicitudes = SolicitudMarca.objects.filter(
        trabajador=request.user,
        estado='PENDIENTE'
    ).order_by('-created_at')

    return render(request, 'asistencia/mis_solicitudes.html', {'solicitudes': solicitudes})

@login_required
def responder_solicitud(request, solicitud_id, accion):
    solicitud = get_object_or_404(SolicitudMarca, id=solicitud_id)
    es_dueno = solicitud.trabajador == request.user
    es_jefe = request.user.is_staff or request.user.is_superuser

    if not (es_dueno or es_jefe):
        messages.error(request, "No tienes permiso para gestionar esta solicitud.")
        return redirect('home')

    if solicitud.estado != 'PENDIENTE':
        messages.error(request, "Esta solicitud ya fue procesada.")
        return redirect('home') # O al panel_rrhh si es jefe

    if accion == 'RECHAZAR':
        solicitud.estado = 'RECHAZADA'
        solicitud.save()
        messages.warning(request, "Has rechazado la corrección. Se generará una incidencia administrativa.")
        # Aquí podrías disparar un email al RRHH avisando del conflicto

    elif accion == 'ACEPTAR':
        try:
            with transaction.atomic():
                # CASO A: NUEVA MARCA
                if solicitud.tipo_solicitud == 'NUEVA':
                    Marcacion.objects.create(
                        trabajador=solicitud.trabajador, # OJO: Usar solicitud.trabajador, no request.user
                        timestamp=solicitud.fecha_hora_propuesta,
                        tipo=solicitud.tipo_marca_propuesta,
                        es_manual=True,
                        observacion=f"Manual por solicitud. Motivo: {solicitud.motivo}"
                    )

                # CASO B: RECTIFICACIÓN
                elif solicitud.tipo_solicitud == 'RECTIFICACION' and solicitud.marca_original:
                    marca_vieja = solicitud.marca_original
                    marca_vieja.estado = 'RECTIFICADA'
                    marca_vieja.save()

                    Marcacion.objects.create(
                        trabajador=solicitud.trabajador,
                        timestamp=solicitud.fecha_hora_propuesta,
                        tipo=solicitud.tipo_marca_propuesta,
                        es_manual=True,
                        marca_reemplazada=marca_vieja,
                        observacion=f"Rectificación aceptada. Motivo: {solicitud.motivo}"
                    )

                solicitud.estado = 'ACEPTADA'
                solicitud.save()
                messages.success(request, "Solicitud aceptada y procesada.")

        except Exception as e:
            messages.error(request, f"Error procesando la solicitud: {str(e)}")

    # --- EL ARREGLO ESTÁ AQUÍ ---
    # Este bloque debe estar FUERA del if/elif de arriba (a la misma altura de indentación)

    if es_jefe and not es_dueno:
        return redirect('panel_rrhh') # Si es el jefe aprobando, vuelve al panel
    else:
        return redirect('home') # Si es el trabajador, vuelve al dashboard

@login_required
def crear_solicitud_trabajador(request):
    if request.method == 'POST':
        tipo_solicitud = request.POST.get('tipo_solicitud')
        fecha_str = request.POST.get('fecha')
        hora_str = request.POST.get('hora')
        motivo = request.POST.get('motivo')

        # Nuevos campos
        tipo_marca = request.POST.get('tipo_marca') # Viene del select manual
        marca_id = request.POST.get('marca_id')     # Viene del select de historial

        # 1. Procesar Fecha y Hora
        try:
            fecha_hora_naive = datetime.strptime(f"{fecha_str} {hora_str}", "%Y-%m-%d %H:%M")
            fecha_hora_aware = timezone.make_aware(fecha_hora_naive)
        except ValueError:
            messages.error(request, "Formato de fecha u hora inválido.")
            return redirect('home')

        # 2. Configurar objeto Solicitud
        solicitud = SolicitudMarca(
            trabajador=request.user,
            solicitante=request.user,
            tipo_solicitud=tipo_solicitud,
            fecha_hora_propuesta=fecha_hora_aware,
            motivo=motivo,
            estado='PENDIENTE'
        )

        # 3. Lógica según el tipo
        if tipo_solicitud == 'RECTIFICACION' and marca_id:
            # Buscamos la marca original que seleccionó el usuario
            try:
                marca_original = Marcacion.objects.get(id=marca_id, trabajador=request.user)
                solicitud.marca_original = marca_original
                # Si es rectificación, el "tipo propuesto" suele ser el mismo de la marca original
                solicitud.tipo_marca_propuesta = marca_original.tipo
            except Marcacion.DoesNotExist:
                messages.error(request, "La marca seleccionada no existe.")
                return redirect('home')
        else:
            # Si es NUEVA, usamos el tipo que eligió manualmente
            solicitud.tipo_marca_propuesta = tipo_marca

        solicitud.save()
        messages.success(request, "Solicitud enviada correctamente.")
        return redirect('home')

    return redirect('home')

@login_required
def exportar_reporte_fiscalizacion(request):
    # 1. Configuración del Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Obtenemos fechas del GET para el nombre del archivo
    desde = request.GET.get('desde', 'inicio')
    hasta = request.GET.get('hasta', 'fin')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Fiscalizacion_DT_{desde}_{hasta}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro de Asistencia"

    # 2. Estilos (Normativa DT)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid") # Azul oscuro
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 3. Definir Columnas
    headers = [
        "ID Registro",
        "RUT Trabajador",
        "Nombre Completo",
        "Fecha",
        "Hora",
        "Tipo de Marca",
        "Origen",
        "Geolocalización",
        "Estado",
        "Checksum (Hash)"
    ]

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 4. Obtener datos
    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.empresa:
        return HttpResponse("Error: Usuario sin empresa asignada.", status=403)

    empresa_actual = perfil.empresa

    # Filtro Base: Empresa
    marcas = Marcacion.objects.filter(
        trabajador__perfil__empresa=empresa_actual
    ).select_related('trabajador', 'trabajador__perfil').order_by('-timestamp')

    # Filtro de Fechas (IMPORTANTE)
    if request.GET.get('desde') and request.GET.get('hasta'):
        desde_fmt = request.GET.get('desde')
        hasta_fmt = request.GET.get('hasta')
        marcas = marcas.filter(timestamp__range=[desde_fmt, hasta_fmt + " 23:59:59"])

    # 5. Llenar filas
    for row_num, marca in enumerate(marcas, 2):
        # A. Preparar datos de Fecha/Hora
        fecha_local = timezone.localtime(marca.timestamp)
        fecha_str = fecha_local.strftime('%d/%m/%Y')
        hora_str = fecha_local.strftime('%H:%M:%S')

        # B. Obtener RUT seguro (CORRECCIÓN CRÍTICA 🛠️)
        # Si existe el perfil y tiene RUT, úsalo. Si no, pon "S/I"
        if hasattr(marca.trabajador, 'perfil') and marca.trabajador.perfil.rut:
            rut_real = marca.trabajador.perfil.rut
        else:
            rut_real = "S/I"

        # C. Generar HASH (Huella digital)
        # Usamos el ID + RUT + Fecha + Hora para que sea único
        raw_data = f"{marca.id}{rut_real}{fecha_str}{hora_str}{marca.tipo}".encode('utf-8')
        hash_seguridad = hashlib.sha256(raw_data).hexdigest()

        # D. Escribir en celdas
        ws.cell(row=row_num, column=1, value=marca.id).alignment = center_align
        ws.cell(row=row_num, column=2, value=rut_real).alignment = center_align  # <--- RUT Real
        ws.cell(row=row_num, column=3, value=marca.trabajador.get_full_name())
        ws.cell(row=row_num, column=4, value=fecha_str).alignment = center_align
        ws.cell(row=row_num, column=5, value=hora_str).alignment = center_align

        # Color según tipo
        cell_tipo = ws.cell(row=row_num, column=6, value=marca.tipo)
        cell_tipo.alignment = center_align
        if marca.tipo == 'ENTRADA':
            cell_tipo.font = Font(color="006600", bold=True)
        elif marca.tipo == 'SALIDA':
            cell_tipo.font = Font(color="990000", bold=True)

        ws.cell(row=row_num, column=7, value="WEB/APP").alignment = center_align

        # Geolocalización limpia
        geo_info = marca.direccion if marca.direccion else f"{marca.latitud}, {marca.longitud}"
        ws.cell(row=row_num, column=8, value=geo_info[:60]) # Recortar para que no rompa el Excel

        # Estado
        estado_texto = "VIGENTE" # Por defecto
        # (Aquí podrías poner lógica si tienes marcas anuladas o rectificadas)
        ws.cell(row=row_num, column=9, value=estado_texto).alignment = center_align

        # Hash (Fuente monoespaciada)
        cell_hash = ws.cell(row=row_num, column=10, value=hash_seguridad)
        cell_hash.font = Font(name='Courier New', size=9, color="555555")

    # 6. Ajustar ancho de columnas manualmente (Más seguro que automático)
    ws.column_dimensions['A'].width = 12 # ID
    ws.column_dimensions['B'].width = 15 # RUT
    ws.column_dimensions['C'].width = 30 # Nombre
    ws.column_dimensions['D'].width = 12 # Fecha
    ws.column_dimensions['E'].width = 12 # Hora
    ws.column_dimensions['H'].width = 40 # Geo
    ws.column_dimensions['J'].width = 65 # Hash

    wb.save(response)
    return response

@login_required
def exportar_reporte_remuneraciones(request):
    # 1. Obtener Empresa y Validar
    perfil_admin = getattr(request.user, 'perfil', None)
    if not perfil_admin or not perfil_admin.empresa:
        return HttpResponse("Error: No tiene empresa asignada.", status=403)

    empresa = perfil_admin.empresa

    # 2. Definir Rango de Fechas
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    if fecha_inicio_str and fecha_fin_str:
        start_date = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    else:
        # Por defecto mes actual
        hoy = timezone.localdate()
        start_date = hoy.replace(day=1)
        # Truco fin de mes
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)

    # 3. Preparar Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Remuneraciones_{empresa.nombre}_{start_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Nomina LRE"

    # Estilos
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    font_white = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal='center')

    # 4. ENCABEZADOS
    headers = [
        "RUT", "Nombre Completo", "Cargo",
        "Días Trab. (Cód 1102)", "Horas Ordinarias",
        "H.E. 50% (Cód 2101)", "H.E. 100% (Dom/Fest)",
        "Min. Atraso", "Días Ausencia (Cód 1106)", "Observaciones"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = font_white
        cell.alignment = align_center

    # 5. OPTIMIZACIÓN: Cargar Feriados y Ausencias en Memoria 🚀
    # Esto evita hacer miles de consultas a la base de datos dentro del bucle
    feriados_set = set(Feriado.objects.filter(fecha__range=[start_date, end_date]).values_list('fecha', flat=True))

    # Trabajadores de la empresa
    trabajadores = User.objects.filter(perfil__empresa=empresa, is_active=True).distinct()

    # 6. ITERACIÓN POR TRABAJADOR
    for trabajador in trabajadores:
        perfil = getattr(trabajador, 'perfil', None)
        if not perfil: continue

        # Configuración personal
        horas_jornada = perfil.jornada_diaria if perfil.jornada_diaria else 9
        hora_entrada_oficial = perfil.hora_entrada if perfil.hora_entrada else time(9,0)

        JORNADA_SEGUNDOS = horas_jornada * 3600

        # Acumuladores
        dias_trabajados = 0
        seg_ordinarios = 0
        seg_extra_50 = 0
        seg_extra_100 = 0
        minutos_atraso = 0
        dias_ausencia = 0
        observaciones = []

        # --- RECORRIDO DÍA A DÍA ---
        delta_days = (end_date - start_date).days + 1

        for i in range(delta_days):
            dia_actual = start_date + timedelta(days=i) # Es objeto date
            dia_semana = dia_actual.weekday() # 0=Lunes ... 6=Domingo

        # Mapa rápido para saber si le toca trabajar este día específico
            mapa_turno = {
                0: perfil.trabaja_lunes,
                1: perfil.trabaja_martes,
                2: perfil.trabaja_miercoles,
                3: perfil.trabaja_jueves,
                4: perfil.trabaja_viernes,
                5: perfil.trabaja_sabado,
                6: perfil.trabaja_domingo,
            }
            le_toca_trabajar = mapa_turno.get(dia_semana, False)

            # A. Revisar si hay justificaciones (Vacaciones / Licencias)
            # Consultamos si HOY cae en un rango de vacación o licencia
            es_vacacion = Vacacion.objects.filter(
                trabajador=trabajador,
                inicio__lte=dia_actual,
                fin__gte=dia_actual,
                estado='APROBADA'
            ).exists()
            es_licencia = LicenciaMedica.objects.filter(trabajador=trabajador, inicio__lte=dia_actual, fin__gte=dia_actual).exists()
            es_feriado = dia_actual in feriados_set
            es_domingo = (dia_actual.weekday() == 6)
            es_sabado = (dia_actual.weekday() == 5)

            # B. Buscar Marcas
            marcas_dia = Marcacion.objects.filter(trabajador=trabajador, timestamp__date=dia_actual)

            if marcas_dia.exists():
                dias_trabajados += 1

                # Calcular Tiempos
                entrada = marcas_dia.filter(tipo='ENTRADA').aggregate(Min('timestamp'))['timestamp__min']
                salida = marcas_dia.filter(tipo='SALIDA').aggregate(Max('timestamp'))['timestamp__max']

                if entrada and salida:
                    duracion = (salida - entrada).total_seconds()

                    # Descuento Colación (1 hora estándar)
                    tiempo_neto = max(0, duracion - 3600)

                    # --- LÓGICA DE HORAS EXTRAS ---
                    # Si es Domingo O Feriado => Todo es al 100%
                    if es_domingo or es_feriado:
                        seg_extra_100 += tiempo_neto
                    else:
                        # Día normal
                        if tiempo_neto > JORNADA_SEGUNDOS:
                            seg_ordinarios += JORNADA_SEGUNDOS
                            seg_extra_50 += (tiempo_neto - JORNADA_SEGUNDOS)
                        else:
                            seg_ordinarios += tiempo_neto

                    # --- CÁLCULO DE ATRASO ---
                    # Solo calculamos atraso si NO es feriado y NO es fin de semana (salvo que tenga turno)
                    if not es_feriado and not es_domingo and not es_sabado:
                        hora_real = timezone.localtime(entrada).time()
                        # Comparación de horas usando datetime dummy
                        dummy = datetime.today().date()
                        dt_real = datetime.combine(dummy, hora_real)
                        dt_oficial = datetime.combine(dummy, hora_entrada_oficial)
                        # Tolerancia 10 min
                        if dt_real > (dt_oficial + timedelta(minutes=10)):
                            diff = dt_real - dt_oficial
                            minutos_atraso += int(diff.total_seconds() / 60)

            else:
                # NO HAY MARCAS (Posible Ausencia)
                # Aquí aplicamos la lógica inteligente:

                if es_feriado:
                    pass # Es feriado, no es ausencia, se paga igual.
                elif es_vacacion:
                    pass # Está de vacaciones pagadas.
                elif es_licencia:
                    pass # Tiene licencia (paga la caja de compensación/Isapre).
                elif not le_toca_trabajar:
                    # Si NO le toca trabajar hoy, NO es ausencia. Es su día libre.
                    pass

                else:
                    # Es día hábil, NO es feriado, LE TOCA trabajar y NO vino => FALTA
                    dias_ausencia += 1

        # Agregar observaciones si tiene muchas faltas
        if dias_ausencia > 0:
            observaciones.append(f"{dias_ausencia} Ausencias injustificadas")
        if es_vacacion: # Solo como dato del último día revisado, mejorable
            observaciones.append("Periodo con Vacaciones")

        # Función auxiliar formateo HH:MM
        def fmt_horas(segundos):
            h = int(segundos // 3600)
            m = int((segundos % 3600) // 60)
            return f"{h:02d}:{m:02d}"

        # 7. ESCRIBIR FILA
        row = [
            perfil.rut,
            f"{trabajador.first_name} {trabajador.last_name}",
            perfil.cargo,
            dias_trabajados,
            fmt_horas(seg_ordinarios),
            fmt_horas(seg_extra_50),
            fmt_horas(seg_extra_100),
            minutos_atraso,
            dias_ausencia,
            ", ".join(observaciones)
        ]
        ws.append(row)

    # Ajuste anchos
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    wb.save(response)
    return response

def privacidad(request):
    return render(request, 'asistencia/privacidad.html')

@csrf_exempt
def mejorar_justificacion_ia(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            texto_informal = data.get('texto', '')

            if len(texto_informal) < 3:
                return JsonResponse({'error': 'El texto es muy corto.'}, status=400)

            # 1. Elegimos el modelo rápido y gratis
            model = genai.GenerativeModel('gemini-1.5-flash')

            # 2. El Prompt (Instrucciones)
            prompt = f"""
            Actúa como un experto en RRHH. Reescribe el siguiente texto informal de un trabajador
            para que sea una justificación de inasistencia/atraso formal y profesional.
            No agregues explicaciones, solo dame el texto corregido.

            Texto informal: "{texto_informal}"
            """

            # 3. Generamos la respuesta
            response = model.generate_content(prompt)

            # Limpiamos un poco el texto por si trae comillas extra
            texto_mejorado = response.text.strip().replace('"', '')

            return JsonResponse({'texto_mejorado': texto_mejorado})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def gestionar_ausencias(request):
    hoy = timezone.localdate()

    # Listas para mostrar en la tabla (Historial)
    vacaciones = Vacacion.objects.all().order_by('-inicio')
    licencias = LicenciaMedica.objects.all().order_by('-inicio')

    # Formularios
    form_vacacion = VacacionForm(request.POST or None)
    form_licencia = LicenciaForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':
        # Detectar qué formulario se envió (usando un campo oculto o nombre del boton)
        if 'btn_vacacion' in request.POST:
            if form_vacacion.is_valid():
                form_vacacion.save()
                messages.success(request, "Vacaciones registradas correctamente.")
                return redirect('gestionar_ausencias')

        elif 'btn_licencia' in request.POST:
            if form_licencia.is_valid():
                form_licencia.save()
                messages.success(request, "Licencia médica registrada correctamente.")
                return redirect('gestionar_ausencias')

    context = {
        'vacaciones': vacaciones,
        'licencias': licencias,
        'form_vacacion': form_vacacion,
        'form_licencia': form_licencia,
        'hoy': hoy
    }
    return render(request, 'asistencia/gestionar_ausencias.html', context)

@login_required
def mis_vacaciones(request):
    usuario = request.user

    # 1. Procesar Solicitud Nueva
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        motivo = request.POST.get('motivo')

        # Validaciones básicas
        if fecha_inicio and fecha_fin:
            # Crear la solicitud en estado PENDIENTE
            Vacacion.objects.create(
                trabajador=usuario,
                inicio=fecha_inicio,
                fin=fecha_fin,
                comentario=motivo,
                estado='PENDIENTE' # Importante: Nace pendiente
            )
            messages.success(request, "🎉 Solicitud enviada a RRHH para aprobación.")
            return redirect('mis_vacaciones')

    # 2. Listar mis vacaciones (Historial)
    mis_solicitudes = Vacacion.objects.filter(trabajador=usuario).order_by('-inicio')

    # Días disponibles (Esto podrías calcularlo real después)
    dias_disponibles = 15

    context = {
        'solicitudes': mis_solicitudes,
        'dias_disponibles': dias_disponibles
    }
    return render(request, 'asistencia/mis_vacaciones.html', context)

@login_required
def aprobar_vacacion(request, id_vacacion, estado):
    # Solo RRHH debería poder hacer esto
    if not request.user.is_staff:
        return redirect('home')

    vacacion = get_object_or_404(Vacacion, id=id_vacacion)
    vacacion.estado = estado
    vacacion.save()

    messages.success(request, f"Solicitud {estado.lower()} correctamente.")
    return redirect('gestionar_ausencias')

@login_required
def mis_dias_administrativos(request):
    # 1. Procesar Solicitud (POST)
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha')
        jornada = request.POST.get('tipo_jornada')
        motivo = request.POST.get('motivo')

        # Guardamos
        DiaAdministrativo.objects.create(
            trabajador=request.user,
            fecha=fecha_str,
            tipo_jornada=jornada,
            motivo=motivo
        )
        messages.success(request, "Solicitud de día administrativo enviada a RRHH.")
        return redirect('mis_dias_administrativos')

    # 2. Listar Historial (GET)
    mis_solicitudes = DiaAdministrativo.objects.filter(trabajador=request.user)

    return render(request, 'asistencia/mis_dias_administrativos.html', {
        'solicitudes': mis_solicitudes
    })

@staff_member_required
def gestionar_dia_administrativo(request, solicitud_id, accion):
    # 1. Buscamos la solicitud (o error 404 si no existe)
    solicitud = get_object_or_404(DiaAdministrativo, id=solicitud_id)

    # 2. Aplicamos la acción
    if accion == 'aprobar':
        solicitud.estado = 'APROBADO'
        messages.success(request, f"Solicitud de {solicitud.trabajador.first_name} Aprobada ✅")

    elif accion == 'rechazar':
        solicitud.estado = 'RECHAZADO'
        messages.error(request, f"Solicitud de {solicitud.trabajador.first_name} Rechazada ❌")

    # 3. Guardamos y volvemos al panel
    solicitud.save()
    return redirect('panel_rrhh')

@login_required
def importar_nomina(request):
    # Solo administradores
    if not request.user.is_superuser:
        messages.error(request, "⛔ Acceso denegado. Esta herramienta es solo para administración técnica.")
        return redirect('home')

    if request.method == 'POST' and request.FILES['archivo_excel']:
        excel_file = request.FILES['archivo_excel']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Contadores
            creados = 0
            actualizados = 0

            # Iteramos desde la fila 2 (saltando encabezados)
            with transaction.atomic(): # Si algo falla grave, no guarda nada a medias
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Asignamos variables según el orden de columnas (0 a 13)
                    # Ajusta estos índices si cambias el orden del Excel
                    if not row[0]: continue # Si no hay username, saltar

                    username = str(row[0]).strip()
                    email = str(row[1]).strip() if row[1] else ""
                    first_name = str(row[2]).strip() if row[2] else ""
                    last_name = str(row[3]).strip() if row[3] else ""
                    rut = str(row[4]).strip() if row[4] else ""
                    cargo = str(row[5]).strip() if row[5] else ""

                    # Hora: Excel a veces devuelve datetime o string
                    hora_entrada = row[6] # Trataremos de procesarlo en el perfil

                    # Función auxiliar para leer SI/NO
                    def es_si(valor):
                        return str(valor).upper().strip() in ['SI', 'S', 'YES', '1', 'TRUE']

                    dias_config = {
                        'lunes': es_si(row[7]),
                        'martes': es_si(row[8]),
                        'miercoles': es_si(row[9]),
                        'jueves': es_si(row[10]),
                        'viernes': es_si(row[11]),
                        'sabado': es_si(row[12]),
                        'domingo': es_si(row[13]),
                    }

                    # 1. Crear o Buscar Usuario
                    user, created = User.objects.get_or_create(username=username)
                    user.email = email
                    user.first_name = first_name
                    user.last_name = last_name

                    if created:
                        # Contraseña por defecto: El RUT (sin puntos ni guion) o '123456'
                        password = rut.replace(".", "").replace("-", "") if rut else "123456"
                        user.set_password(password)
                        creados += 1
                    else:
                        actualizados += 1

                    user.save()

                    # 2. Configurar Perfil
                    perfil, _ = Perfil.objects.get_or_create(usuario=user)
                    perfil.rut = rut
                    perfil.cargo = cargo

                    # Asignar Empresa (Asumimos la del admin que carga)
                    if request.user.perfil.empresa:
                        perfil.empresa = request.user.perfil.empresa

                    # Asignar Hora Entrada
                    if hora_entrada:
                        import datetime
                        if isinstance(hora_entrada, datetime.time):
                            perfil.hora_entrada = hora_entrada
                        elif isinstance(hora_entrada, datetime.datetime):
                            perfil.hora_entrada = hora_entrada.time()
                        elif isinstance(hora_entrada, str):
                            # Intentar convertir string "09:00"
                            try:
                                h, m = map(int, hora_entrada.split(':'))
                                perfil.hora_entrada = datetime.time(h, m)
                            except:
                                pass # Si falla, queda la default

                    # 3. Asignar Días de Trabajo (Lo nuevo)
                    perfil.trabaja_lunes = dias_config['lunes']
                    perfil.trabaja_martes = dias_config['martes']
                    perfil.trabaja_miercoles = dias_config['miercoles']
                    perfil.trabaja_jueves = dias_config['jueves']
                    perfil.trabaja_viernes = dias_config['viernes']
                    perfil.trabaja_sabado = dias_config['sabado']
                    perfil.trabaja_domingo = dias_config['domingo']

                    perfil.save()

            messages.success(request, f"Proceso terminado: {creados} nuevos, {actualizados} actualizados.")

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return render(request, 'asistencia/importar_nomina.html')